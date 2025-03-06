from symtable import Class

import openpyxl
# from openpyxl.styles import Color
# from openpyxl.worksheet.print_settings import PrintTitles


# Очень большой класс для такой малой задачи


class Config:

    def __init__(self):
        self.Ai = dict()
        self.Di = dict()
        self.Ao = dict()
        self.Do = dict()
        self.RegUi = dict()
        self.RegUo = dict()
        self.Var = dict()
        self.typeio = dict()
        self.Method = dict()




    def readExel(self, name_file):# Не замечание лишь рекомендация:  имена методов / функции начинаются с маленькой буквы
                       # https://peps.python.org/pep-0008/ или https://python.ivan-shamaev.ru/pep8-python-code-rules-programmers-guide/

        wb = openpyxl.load_workbook(name_file) # плохое название файла)


        for item in self.__dict__.keys():
            dict_temp = dict()
            sheet = wb[item]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                data_tuple = row  # row уже будет кортежем
                dict_temp[data_tuple[0]] = (data_tuple[1:])
            setattr(self, item, dict_temp)

    def getAllConfig(self):
        return self




    def printDictionary(self):
        pass
        for item in self.__dict__.values(): # из той же оперы как вывести для отладки
            print(item)
        print('='*100)




